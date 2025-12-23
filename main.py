# Write your code here and merge with main branch 
import pandas as pd

# Read Excel file
df = pd.read_excel("employees.xlsx")

# Show first rows
print(df.head())

# Access columns
print(df["Name"])

# Iterate through rows
for index, row in df.iterrows():
    print(row["ID"], row["Name"], row["Department"], row["Salary"])





# Parse Multiple Sheets from One Excel File

import pandas as pd

# Load all sheets into a dictionary
sheets = pd.read_excel("employees.xlsx", sheet_name=None)

for sheet_name, df in sheets.items():
    print(f"\nSheet: {sheet_name}")
    print(df)





#Parse Excel Using openpyxl (Without Pandas)

from openpyxl import load_workbook

# Load workbook
wb = load_workbook("employees.xlsx")

# Select sheet
ws = wb["Employees"]

# Iterate through rows
for row in ws.iter_rows(min_row=2, values_only=True):
    emp_id, name, department, salary = row
    print(emp_id, name, department, salary)




##Convert Excel

import pandas as pd

df = pd.read_excel("employees.xlsx")

records = df.to_dict(orient="records")
print(records)


